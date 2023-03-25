import numpy as np
import openpyxl as opxl
import pandas as pd

#load workbook
workbook = opxl.load_workbook(filename="F:\\FinancialsSampleData.xlsx", data_only=True)

def get_sheet_names(wb):
    sheet_names = wb.get_sheet_names()
    sheet1 = wb.get_sheet_by_name('Financials')
    sheet1_title = sheet1.title

    # another_sheet =  wb.get_active_sheet()
    # sheet2_title = another_sheet.title
    return sheet1_title


def get_cell_values(wb):
    sheet = wb.get_sheet_by_name('Financials')

    #sheet1[A1] --> will give us the cell referecnce
    #sheet1[A1].value --> will give the cell value

    Account = sheet['A3'].value
    Unit = sheet['B3'].value
    Year = sheet['D3'].value
    Actual = sheet['F3'].value

    return 'From ' + Account + ' Under ' + Unit + ' In ' + str(Year) + ' The actual value is ' + str(Actual)


def get_cell_with_row_nd_column_num(wb):
    sheet = wb.get_sheet_by_name('Financials')

    cell_reference = sheet.cell(row=3, column=1)
    cell_value = sheet.cell(row=3, column=1).value

    cell_values = []
    # looping using cell reference numbers
    for i in range(1, 10, 3):
        cell_values.append(sheet.cell(row = i, column = 2).value)
    
    return cell_values

def get_the_size_of_the_sheet(wb):
    sheet = wb.get_sheet_by_name('Financials')

    highest_row = sheet.get_highest_row()
    highest_column = sheet.get_highest_column()

    return highest_row, highest_column


